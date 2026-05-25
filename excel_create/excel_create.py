import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def export_to_excel():
    print("Lendo dados do CSV...")
    # Lendo o CSV mapeando o ponto e vírgula correto
    df = pd.read_csv('csv/produtos.csv', sep=';', decimal=',')
    
    # ----------------------------------------------------
    # ATENÇÃO: Substitua aqui pelos nomes exatos das colunas do seu CSV
    col_produto = 'Produto'
    col_preco = 'Preco'
    col_estoque = 'Qtde_Disponivel'
    # ----------------------------------------------------

    # Inicializando o Workbook do Excel
    wb = openpyxl.Workbook()
    
    # Configurando as Abas
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_data = wb.create_sheet(title="Dados do Estoque")
    
    # Garantindo que as linhas de grade fiquem visíveis
    ws_dash.views.sheetView[0].showGridLines = True
    ws_data.views.sheetView[0].showGridLines = True

    # Estilos Visuais (Paleta Azul Corporativa/Cool Slate)
    font_title = Font(name="Segoe UI", size=16, bold=True, color="1F497D")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=11, color="000000")
    font_bold = Font(name="Segoe UI", size=11, bold=True, color="000000")
    font_kpi_val = Font(name="Segoe UI", size=18, bold=True, color="1F497D")
    font_kpi_lbl = Font(name="Segoe UI", size=9, bold=True, color="595959")

    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    fill_kpi = PatternFill(start_color="E9EEF4", end_color="E9EEF4", fill_type="solid")
    fill_total = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    thin_side = Side(border_style="thin", color="D9D9D9")
    border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_total = Border(top=Side(border_style="thin", color="000000"), bottom=Side(border_style="double", color="000000"))

    # --- ABA 2: CONSTRUÇÃO DA TABELA DE DADOS ---
    headers = ["Produto", "Preço Unitário", "Qtd em Estoque", "Valor em Estoque"]
    ws_data.append([]) # Linha em branco inicial
    ws_data.append(headers)

    # Estilizando o cabeçalho
    for cell in ws_data[2]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Inserindo as linhas dinamicamente
    for i, row in df.iterrows():
        r_idx = i + 3
        ws_data.cell(row=r_idx, column=1, value=row[col_produto]).alignment = Alignment(horizontal="left")
        ws_data.cell(row=r_idx, column=2, value=float(row[col_preco])).number_format = 'R$ #,##0.00'
        ws_data.cell(row=r_idx, column=3, value=int(row[col_estoque])).number_format = '#,##0'
        
        # Fórmula nativa multiplicando as células correspondentes (Preço * Quantidade)
        ws_data.cell(row=r_idx, column=4, value=f"=B{r_idx}*C{r_idx}").number_format = 'R$ #,##0.00'
        
        # Aplicando fontes, bordas e listras alternadas (Zebra)
        for c in range(1, 5):
            cell = ws_data.cell(row=r_idx, column=c)
            cell.font = font_body
            cell.border = border_cell
            if i % 2 == 1:
                cell.fill = fill_zebra

    # Linha de Totais da tabela
    total_idx = len(df) + 3
    ws_data.cell(row=total_idx, column=1, value="Total Geral").font = font_bold
    ws_data.cell(row=total_idx, column=3, value=f"=SUM(C3:C{total_idx-1})").number_format = '#,##0'
    ws_data.cell(row=total_idx, column=4, value=f"=SUM(D3:D{total_idx-1})").number_format = 'R$ #,##0.00'

    for c in range(1, 5):
        cell = ws_data.cell(row=total_idx, column=c)
        cell.border = border_total
        cell.fill = fill_total
        if c > 1:
            cell.alignment = Alignment(horizontal="right" if c != 3 else "center")

    # Ajuste automático do tamanho das colunas baseado no conteúdo
    for col in ws_data.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_data.column_dimensions[col_letter].width = max(max_len + 4, 15)


    # --- ABA 1: CONSTRUÇÃO DO DASHBOARD ---
    ws_dash.cell(row=2, column=2, value="Relatório Gerencial de Estoque").font = font_title
    
    # Bloco KPI 1: Variedade de Itens
    ws_dash.merge_cells("B4:C4"); ws_dash.merge_cells("B5:C5")
    ws_dash.cell(row=4, column=2, value="PRODUTOS CADASTRADOS").font = font_kpi_lbl
    ws_dash.cell(row=5, column=2, value=len(df)).font = font_kpi_val
    
    # Bloco KPI 2: Total de Unidades (Buscando dinamicamente do total da aba de dados)
    ws_dash.merge_cells("E4:F4"); ws_dash.merge_cells("E5:F5")
    ws_dash.cell(row=4, column=5, value="TOTAL DE UNIDADES").font = font_kpi_lbl
    ws_dash.cell(row=5, column=5, value=f"='Dados do Estoque'!C{total_idx}").number_format = '#,##0'
    ws_dash.cell(row=5, column=5).font = font_kpi_val

    # Bloco KPI 3: Valor Patrimonial Total do Estoque
    ws_dash.merge_cells("H4:I4"); ws_dash.merge_cells("H5:I5")
    ws_dash.cell(row=4, column=8, value="VALOR DO FINANCEIRO").font = font_kpi_lbl
    ws_dash.cell(row=5, column=8, value=f"='Dados do Estoque'!D{total_idx}").number_format = 'R$ #,##0.00'
    ws_dash.cell(row=5, column=8).font = font_kpi_val

    # Formatando o visual das caixas do Dashboard
    for r in [4, 5]:
        for c in [2, 3, 5, 6, 8, 9]:
            cell = ws_dash.cell(row=r, column=c)
            cell.fill = fill_kpi
            cell.border = border_cell
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Dimensionando as colunas do Dashboard para ficarem simétricas
    for c in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        ws_dash.column_dimensions[c].width = 13

    # Salvando planilha
    nome_saida = "excel_create/relatorio_estoque.xlsx"
    wb.save(nome_saida)
    print(f"Planilha exportada com sucesso em: {nome_saida}")